from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
import os
import shutil

def extract_images_and_positions(sheet):
    images_info = []
    for idx, image in enumerate(sheet._images):
        # Obtener la posición de la imagen (celda) y su tamaño
        image_data = {
            "name": f"image_{idx}.png",  # Generar un nombre único para la imagen
            "anchor": image.anchor,  # La referencia de la celda donde se inserta la imagen
            "width": image.width,  # Ancho de la imagen
            "height": image.height,  # Alto de la imagen
            "ref": image.ref  # Referencia de la imagen dentro del archivo Excel (BytesIO)
        }
        images_info.append(image_data)
    return images_info

def copy_images_between_workbooks(src_path, dest_path):
    # Crear directorio temporal para las imágenes
    temp_dir = "temp_images"
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)

    # Cargar el libro de origen
    wb_src = load_workbook(src_path)

    # Crear un nuevo libro de destino
    wb_dest = Workbook()
    sheet = wb_dest.active
    sheet.title = 'Imágenes Encontradas'

    for sheet_name in wb_src.sheetnames:
        src_sheet = wb_src[sheet_name]
        
        # Obtener imágenes y posiciones de la hoja de origen
        images_info = extract_images_and_positions(src_sheet)

        for image_info in images_info:
            # Guardar la imagen con un nombre único en un archivo temporal
            temp_image_path = os.path.join(temp_dir, image_info['name'])
            with open(temp_image_path, "wb") as f:
                f.write(image_info["ref"].getbuffer())

            # Crear una nueva imagen y asignar tamaño y posición
            img = Image(temp_image_path)
            img.width = image_info["width"]
            img.height = image_info["height"]
            sheet.add_image(img, image_info["anchor"])

    # Guardar el libro de destino
    wb_dest.save(dest_path)

    # Eliminar el directorio temporal y todo su contenido
    shutil.rmtree(temp_dir)

# Rutas de los archivos de origen y destino
src_excel = "C:\\Users\\javier.puentes\\ser_excel\\SIIF_IDEA\\PLANTILLAS\\PSRH2060.xlsx"
dest_excel = "C:\\Users\\javier.puentes\\ser_excel\\SIIF_IDEA\\PLANTILLAS\\ruta_al_archivo_destino.xlsx"

# Copiar imágenes entre libros
copy_images_between_workbooks(src_excel, dest_excel)
