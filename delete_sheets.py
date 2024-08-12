from openpyxl import load_workbook


def delete_sheets(file_path, sheet_principal):
    try:
        wb = load_workbook(file_path)
        todas_las_hojas = wb.sheetnames

        if sheet_principal not in todas_las_hojas:
            raise ValueError(f"La hoja principal '{sheet_principal}' no existe en el archivo.")

        for hoja in todas_las_hojas:
            if hoja != sheet_principal:
                print(f"Eliminando hoja: {hoja}")
                del wb[hoja]

        wb.save(file_path)
    except Exception as e:
        print(f"Error: {e}")