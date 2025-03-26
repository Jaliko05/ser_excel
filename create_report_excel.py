import io
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
import win32com.client as win32
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
import os
import re
import math
import copy
from uuid import uuid4

from generate_barcode import generate_barcode
from log import log
from openpyxl.cell.cell import MergedCell

def obtener_info_excel(ruta_excel):
    workbook = load_workbook(ruta_excel)
    info_excel = {}

    for sheet in workbook.worksheets:
        print("obtener_info_hoja: ",sheet.title)
        sheet_info = {}
        # Obtener el ancho de las columnas y la altura de las filas
        column_widths = {col: sheet.column_dimensions[col].width for col in sheet.column_dimensions}
        row_heights = {row: sheet.row_dimensions[row].height for row in sheet.row_dimensions}

        # Inicializamos una bandera para saber hasta qué fila debemos tomar la información
        fila_limite = None

        for row in sheet.iter_rows():
            # Verificar si hay un valor en la columna A de esta fila
            if row[0].value is not None:
                fila_limite = row[0].row  # Guardar la fila donde se encuentra el valor en la columna A

            # Si hemos encontrado un valor en la columna A, procesar hasta esa fila
            if fila_limite is not None and row[0].row > fila_limite:
                break  # Salir del bucle una vez que se supera la fila con el valor en la columna A

            for cell in row:
                # Omitir celdas fusionadas que no son las principales
                if isinstance(cell, MergedCell):
                    continue

                # Obtener propiedades de estilo, incluso si la celda está vacía
                font_color = cell.font.color.rgb if cell.font.color and cell.font.color.type == 'rgb' else None
                fill_color = cell.fill.fgColor.rgb if cell.fill.fgColor and cell.fill.fgColor.type == 'rgb' else None
                border_color = cell.border.left.color.rgb if cell.border.left.color and cell.border.left.color.type == 'rgb' else None
                border_styles = {
    'left': {'style': cell.border.left.style, 'color': cell.border.left.color.rgb if cell.border.left.color and hasattr(cell.border.left.color, 'rgb') else None},
    'right': {'style': cell.border.right.style, 'color': cell.border.right.color.rgb if cell.border.right.color and hasattr(cell.border.right.color, 'rgb') else None},
    'top': {'style': cell.border.top.style, 'color': cell.border.top.color.rgb if cell.border.top.color and hasattr(cell.border.top.color, 'rgb') else None},
    'bottom': {'style': cell.border.bottom.style, 'color': cell.border.bottom.color.rgb if cell.border.bottom.color and hasattr(cell.border.bottom.color, 'rgb') else None}
                }


                # Verificar si hay algún estilo modificado (bordes, color de fuente, color de fondo)
                has_styles = any([
                    font_color,  # Color de fuente
                    fill_color,  # Color de relleno
                    border_color,
                    border_styles['left'],  # Bordes
                    border_styles['right'],
                    border_styles['top'],
                    border_styles['bottom']
                ])

                # Considerar la celda solo si tiene un valor o estilos aplicados (ignorar si solo tiene cambios de tamaño)
                if cell.value is not None or has_styles:
                    # Obtener la configuración de la página
                    page_setup = {
                        'orientation': sheet.page_setup.orientation,
                        'paper_size': sheet.page_setup.paperSize,
                        'fit_to_width': sheet.page_setup.fitToWidth,
                        'fit_to_height': sheet.page_setup.fitToHeight,
                        'scale': sheet.page_setup.scale,
                        'margin_top': sheet.page_margins.top,
                        'margin_bottom': sheet.page_margins.bottom,
                        'margin_left': sheet.page_margins.left,
                        'margin_right': sheet.page_margins.right,
                        'print_area': sheet.print_area
                    }

                    # Obtener el ancho de la columna y la altura de la fila desde `column_dimensions` y `row_dimensions`
                    col_letter = cell.column_letter
                    row_number = cell.row
                    column_width = column_widths.get(col_letter, None)
                    row_height = row_heights.get(row_number, None)

                    text_rotation = cell.alignment.textRotation

                    cell_info = {
                        'value': cell.value,
                        'font': {
                            'name': cell.font.name,
                            'size': cell.font.sz,
                            'bold': cell.font.b,
                            'italic': cell.font.i,
                            'underline': cell.font.u,
                            'color': font_color
                        },
                        'fill': {
                            'fgColor': fill_color,
                            'patternType': cell.fill.patternType
                        },
                        'border': border_styles,
                        'alignment': {
                            'horizontal': cell.alignment.horizontal,
                            'vertical': cell.alignment.vertical,
                            'wrap_text': cell.alignment.wrap_text,
                            'text_rotation': text_rotation
                        },
                        'number_format': cell.number_format,
                        'row': row_number,
                        'column': col_letter,
                        'merge_cells': cell.coordinate in sheet.merged_cells,
                        'column_width': column_width,
                        'row_height': row_height
                    }
                    sheet_info[cell.coordinate] = cell_info

        merge_info = []
        for merged_cell in sheet.merged_cells.ranges:
            merge_info.append(merged_cell)

        info_excel[sheet.title] = {'cells': sheet_info, 'merges': merge_info, 'page_setup': page_setup}

    return info_excel

def es_numero(valor):
    # Verificar si el valor es un número que se pueda convertir a float
    try:
        # Intentar convertir a float, reemplazando coma por punto si es necesario
        float(valor.replace(',', '.'))
        return True
    except ValueError:
        return False

def reemplazar_vars(sheet_info, data, ruta_imagenes='img_barcode'):
    # Hacer una copia profunda del sheet_info original
    sheet_info_copia = copy.deepcopy(sheet_info)

    # Crear la carpeta de imágenes si no existe
    if not os.path.exists(ruta_imagenes):
        os.makedirs(ruta_imagenes)

    for var_counter, value in enumerate(data, start=1):
        var_placeholder = f'<VAR{var_counter:03}>'
        barcode_placeholder = f'<CB{var_counter:03}>'

        # Reemplazar valores en las celdas
        for cell_info in sheet_info_copia['cells'].values():
            if isinstance(cell_info['value'], str):
                # Si el placeholder está en el valor de la celda
                if var_placeholder in cell_info['value']:
                    # Reemplazar el valor en la celda
                    cell_info['value'] = cell_info['value'].replace(var_placeholder, str(value))

                # Generar y reemplazar código de barras
                if barcode_placeholder in cell_info['value']:
                    uuid = str(uuid4())
                    nombre_imagen = f"{ruta_imagenes}/barcode_{uuid}"
                    generate_barcode(str(value), nombre_imagen)
                    cell_info['value'] = cell_info['value'].replace(barcode_placeholder, nombre_imagen + '.png')

    return sheet_info_copia


from openpyxl.drawing.image import Image

def aplicar_info_a_hoja(sheet, sheet_info, start_row, sheet_template):
    max_row = start_row
    nameImge = []	
    

    def convertir_valor_segun_formato(value, number_format):
        if value is None:
            return None  # Si el valor es None, no convertirlo

        # Detectar si el formato es numérico o de moneda
        is_numeric_format = re.match(r'^[0#,.]*[0#]$', number_format) or number_format.lower() == 'general'
        is_currency_format = re.search(r'[\$€¥]', number_format)

        # Si el formato es numérico o de moneda, procesar para quitar decimales y ceros a la izquierda
        if is_numeric_format or is_currency_format:
            try:
                # Quitar ceros a la izquierda al convertir a entero
                value = str(value).lstrip('0') or '0'

            except ValueError:
                return value  # Si no es un valor numérico, continúa sin modificarlo


        # Detectar cualquier formato que sea numérico
        if is_numeric_format:
            try:
                if isinstance(value, str):
                    # Caso 1: Si el valor tiene ambos ',' y '.' (ej: "59,400.00" → "59400.00")
                    if ',' in value and '.' in value:
                        if value.index(',') < value.index('.'):  # Caso "59,400.00"
                            value = value.replace(',', '')  # Eliminar comas, deja el punto decimal
                            value = float(value) 
                        else:  # Caso "59.400,00"
                            value = value.replace('.', '').replace(',', '.')  # Corrige separadores
                            value = float(value) 
                        
                    # Caso 2: Si el valor solo tiene comas, convertir a formato correcto (ej: "59,400" → "59.400")
                    elif ',' in value:
                        value = value.replace(',', '')
                        value = int(value)

                    # Caso 3: Si el valor solo tiene puntos y no es decimal (ej: "59.400" → "59.400")
                    elif '.' in value:
                        value = value.replace('.', '')
                        value = int(value)
                    else:
                        value = int(value)

            except ValueError:
                return value

        # Detectar cualquier formato que sea de moneda
        elif is_currency_format:
            try:
                if isinstance(value, str):
                    # Caso 1: Si el valor tiene ambos ',' y '.' (ej: "59,400.00" → "59400.00")
                    if ',' in value and '.' in value:
                        if value.index(',') < value.index('.'):  # Caso "59,400.00"
                            value = value.replace(',', '')  # Eliminar comas, deja el punto decimal
                            value = float(value) 
                        else:  # Caso "59.400,00"
                            value = value.replace('.', '').replace(',', '.')  # Corrige separadores
                            value = float(value) 
                        
                    # Caso 2: Si el valor solo tiene comas, convertir a formato correcto (ej: "59,400" → "59.400")
                    elif ',' in value:
                        value = value.replace(',', '')
                        value = int(value)

                    # Caso 3: Si el valor solo tiene puntos y no es decimal (ej: "59.400" → "59.400")
                    elif '.' in value:
                        value = value.replace('.', '')
                        value = int(value)
                    else:
                        value = int(value)

            except ValueError:
                return value

        # Detectar cualquier formato que sea de porcentaje
        elif '%' in number_format:
            try:
                value = float(value.replace(',', '')) if isinstance(value, str) else float(value)
                return "{:.2%}".format(value / 100)
            except ValueError:
                return value

        # Detectar formatos de fecha (se basa en los patrones comunes de fechas en Excel)
        elif re.search(r'(m|d|y|M|D|Y)', number_format):
            try:
                return value.strftime("%Y-%m-%d")  # Ajustar según el formato de fecha
            except (ValueError, AttributeError):
                return value

        # Si no es un formato numérico, moneda o conocido, devolver el valor tal cual
        return value


    for coord, cell_info in sheet_info['cells'].items():
        col_letter = ''.join(filter(str.isalpha, coord))
        row_number = int(''.join(filter(str.isdigit, coord)))
        new_coord = f"{col_letter}{start_row + row_number - 1}"

        # Verificar si la celda es parte de un rango fusionado
        is_main_merged_cell = False
        if any(new_coord in range_obj for range_obj in sheet.merged_cells.ranges):
            if any(new_coord == range_obj.coord.split(":")[0] for range_obj in sheet.merged_cells.ranges):
                is_main_merged_cell = True
            else:
                # Considerar la fila incluso si no es la celda principal
                max_row = max(max_row, start_row + row_number - 1)
                continue
        else:
            is_main_merged_cell = True

        # Actualizar la fila máxima considerando todas las filas
        max_row = max(max_row, start_row + row_number - 1)

        # Aplicar la altura de la fila
        if 'row_height' in cell_info and cell_info['row_height'] is not None:
            sheet.row_dimensions[start_row + row_number - 1].height = cell_info['row_height']

        if col_letter != 'A'and is_main_merged_cell:  # Evitar escribir en la columna A
            cell = sheet[new_coord]

            # Verificar si el valor es una URL de una imagen de código de barras
            if isinstance(cell_info['value'], str) and cell_info['value'].endswith('.png'):
                # Crear el objeto de la imagen
                img = Image(cell_info['value'])

                # Usar la función ajustar_imagen_a_celda para ajustar el tamaño de la imagen
                img_info = {'col': column_index_from_string(col_letter), 'row': row_number}
                img = ajustar_imagen_a_celda(sheet_template, img_info, img, 0)

                heigth = img.height * 1.8
                img.height = heigth
                # Insertar la imagen en la celda ajustada
                img.anchor = new_coord  # Posicionar la imagen en la celda correspondiente
                sheet.add_image(img)
                nameImge.append(cell_info['value'])
            else:
                # Convertir el valor según el formato de la celda
                valor_convertido = convertir_valor_segun_formato(cell_info['value'], cell_info['number_format'])
                cell.value = valor_convertido
                
                # Aplicar los estilos
                cell.font = Font(
                    name=cell_info['font']['name'],
                    size=cell_info['font']['size'],
                    bold=cell_info['font']['bold'],
                    italic=cell_info['font']['italic'],
                    underline=cell_info['font']['underline'],
                    color=Color(rgb=cell_info['font']['color']) if cell_info['font']['color'] else None
                )
                cell.fill = PatternFill(
                    fgColor=Color(rgb=cell_info['fill']['fgColor']) if cell_info['fill']['fgColor'] else "FFFFFF", patternType=cell_info['fill']['patternType']
                )

                cell.border = Border(
    left=Side(
        style=cell_info['border']['left']['style'],
        color=Color(rgb=cell_info['border']['left']['color']) if isinstance(cell_info['border']['left']['color'], str) else None
    ),
    right=Side(
        style=cell_info['border']['right']['style'],
        color=Color(rgb=cell_info['border']['right']['color']) if isinstance(cell_info['border']['right']['color'], str) else None
    ),
    top=Side(
        style=cell_info['border']['top']['style'],
        color=Color(rgb=cell_info['border']['top']['color']) if isinstance(cell_info['border']['top']['color'], str) else None
    ),
    bottom=Side(
        style=cell_info['border']['bottom']['style'],
        color=Color(rgb=cell_info['border']['bottom']['color']) if isinstance(cell_info['border']['bottom']['color'], str) else None
    )
)
                cell.alignment = Alignment(
                    horizontal=cell_info['alignment']['horizontal'],
                    vertical=cell_info['alignment']['vertical'],
                    wrap_text=cell_info['alignment']['wrap_text'],
                    text_rotation=cell_info['alignment']['text_rotation']
                )
                cell.number_format = cell_info['number_format']

    # Aplicar los rangos de celdas fusionadas
    for merge_range in sheet_info['merges']:
        min_col, min_row, max_col, max_row_fin = merge_range.bounds

        new_merge_start = f"{get_column_letter(min_col)}{start_row + min_row - 1}"
        new_merge_end = f"{get_column_letter(max_col)}{start_row + max_row_fin - 1}"
        sheet.merge_cells(f"{new_merge_start}:{new_merge_end}")

    return max_row, nameImge

def find_next_start_row(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == '??FIN??':
                return cell.row + 1
            elif cell.value == "0":
                return cell.row + 1
    return 1

import tempfile
def get_image_position_openpyxl(rout_template_excel):
    try:
        wb = load_workbook(rout_template_excel)
        posiciones_imagenes = {}
        temp_dir = tempfile.mkdtemp()
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            posiciones_imagenes[sheet_name] = []
            
            for idx, image in enumerate(sheet._images):
                temp_image_path = os.path.join(temp_dir, f"temp_image_{sheet_name}_{idx}.png")
                
                with open(temp_image_path, 'wb') as img_file:
                    img_file.write(image._data())
                
                # Corregir el manejo de la columna
                original_col = image.anchor._from.col
                if isinstance(original_col, str):
                    col = column_index_from_string(original_col)
                else:
                    col = original_col
                
                posiciones_imagenes[sheet_name].append({
                    "temp_path": temp_image_path,
                    "col": col + 1,  # Ajuste final de índice
                    "row": image.anchor._from.row + 1,
                    "width": image.width,
                    "height": image.height
                })
        print("posiciones_imagenes: ", posiciones_imagenes)
        return posiciones_imagenes
    finally:
        wb.close()

from openpyxl.utils import get_column_letter

def copy_column_widths(origen, destino):
    # Obtener la última columna con un ancho modificado
    last_col_index = max([column_index_from_string(col) for col in origen.column_dimensions if origen.column_dimensions[col].width is not None])

    # Iterar secuencialmente desde la primera columna hasta la última columna modificada
    for col in range(1, last_col_index + 1):
        col_letter = get_column_letter(col)
        origen_ancho = origen.column_dimensions[col_letter].width
        # Mostrar el ancho obtenido
        if origen_ancho - 0.5 > 0:
            origen_ancho = origen_ancho - 0
        # Aplicar el ancho a la columna en la hoja de destino
        print("ancho", origen_ancho)
        destino.column_dimensions[col_letter].width = origen_ancho


def obtener_posicion_celda(img_info, start_row):
    # Convertir número de columna a letra si es necesario
    if isinstance(img_info['col'], int):
        col_letter = get_column_letter(img_info['col'])
    else:
        col_letter = str(img_info['col'])
    
    return f"{col_letter}{img_info['row'] + start_row}"

def obtener_area_celda_combinada(sheet, col_letter, row):
    for merged_cells in sheet.merged_cells.ranges:
        # Obtener los límites del rango combinado
        min_col, min_row, max_col, max_row = merged_cells.bounds
        # Verificar si la celda está dentro de estos límites
        if (min_row <= row <= max_row) and (min_col <= column_index_from_string(col_letter) <= max_col):
            return merged_cells
    return None

def ajustar_imagen_a_celda(sheet, img_info, new_image, start_row):
    col_letter = get_column_letter(img_info['col'])
    row = img_info['row']
    # Verificar si la celda está combinada
    merged_range = obtener_area_celda_combinada(sheet, col_letter, row)

    if merged_range:
        min_col, min_row, max_col, max_row = merged_range.bounds
        col_width = sum(sheet.column_dimensions[get_column_letter(c)].width or 8.43 for c in range(min_col, max_col + 1))
        print(f"Filas combinadas: {list(range(min_row, max_row + 1))}")
        for r in range(min_row, max_row + 1):
            print(f"Altura de fila {r}: {sheet.row_dimensions[r].height}")
        row_height = sum(sheet.row_dimensions[r].height if sheet.row_dimensions[r].height else 15 for r in range(min_row, max_row + 1))
    else:
        col_width = sheet.column_dimensions[col_letter].width or 8.43
        row_height = sheet.row_dimensions[row].height or 15



    # Conversiones aproximadas:
    pixel_width = col_width * 7
    pixel_height = row_height * 1.2  # Ajusta este valor si es necesario


    # Ajustar el tamaño de la imagen al tamaño de la celda o celdas combinadas
    new_image.width = pixel_width
    new_image.height = pixel_height

    return new_image

def aplicar_imagenes_a_hoja(sheet, posiciones_imagenes, template_sheet, start_row):
    for img_info in posiciones_imagenes:
        try:
            # Cargar desde archivo temporal
            img = Image(img_info['temp_path'])
            
            # Ajustar tamaño
            img = ajustar_imagen_a_celda(template_sheet, img_info, img, start_row)
            
            # Posicionar
            cell_position = obtener_posicion_celda(img_info, start_row - 1)
            print(f"Insertando imagen en {cell_position} ({img_info['temp_path']})")
            sheet.add_image(img, cell_position)
            
        except Exception as e:
            print(f"Error insertando imagen: {str(e)}")
            continue

def create_report_excel(datos_report, ruta_template_excel, ruta_report_excel, rout_log):
    message = "Inicio de la copia del archivo: " + ruta_template_excel + "\n"
    try: 
        wb = load_workbook(ruta_template_excel)

        # Buscar la hoja "PRINCIPAL" o variantes, 
        principal_sheet = None
        for sheet_name in ["PRINCIPAL", "principal", "Principal"]:
            if sheet_name in wb.sheetnames:
                principal_sheet = wb[sheet_name]
                break
        if not principal_sheet:
            principal_sheet = wb.create_sheet(title="PRINCIPAL")
        message = message + "Hoja principal: " + str(principal_sheet) + "\n"

        # Obtener la fila de inicio
        start_row = find_next_start_row(principal_sheet)
        message = message + "Obtener fila inicial exitosamente: "  + "\n"
        print("inicio fila: ", start_row)


        # Obtener la información de cada hoja de la plantilla
        info_excel = obtener_info_excel(ruta_template_excel)
        print("obtener_info_excel: ")
        message = message + "Obtener informacion de plantilla exitosamente: "  + "\n"

        #Obtener las posiciones de las imagenes
        posiciones_imagenes = get_image_position_openpyxl(ruta_template_excel)
        message = message + "Obtener posiciones de las imagenes exitosamente: "  + "\n"
        
        #iniciar libro de excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = principal_sheet.title
        
        #copiar ancho de columnas de la hoja 001 al reporte
        copy_column_widths(wb['001'], sheet)
        message = message + "Copiar ancho de columnas de la hoja 001 al reporte exitosamente: "  + "\n"

        #aplicar informacion de la hoja principal
        start_row, imag = aplicar_info_a_hoja(sheet, info_excel[principal_sheet.title], 1, principal_sheet.title)   
        print("inicio fila: ", start_row)
        message = message + "Aplicar informacion de la hoja principal exitosamente: "  + "\n"

        #aplicar imagenes de la hoja principal a la nueva hoja principal
        if posiciones_imagenes[principal_sheet.title]:
            sheet_template = wb[principal_sheet.title]
            aplicar_imagenes_a_hoja(sheet, posiciones_imagenes[sheet_name], sheet_template, start_row)
            message = message + "Aplicar imagenes de la hoja principal exitosamente: "  + "\n"
            print("aplicar imagenes de la hoja principal exitosamente")

        bar_code = []

        # Iterar sobre los datos de reporte y las hojas correspondientes
        for data in datos_report:
            for sheet_name, values in data.items():
                if sheet_name in info_excel:
                    sheet_info = info_excel[sheet_name]
                    # Reemplazar las variables en la hoja
                    sheet_info_modificada = reemplazar_vars(sheet_info, values)
                    # Aplicar la información modificada a la hoja "PRINCIPAL"
                    sheet_template = wb[sheet_name]
                    max_row, nameImge = aplicar_info_a_hoja(sheet, sheet_info_modificada, start_row, sheet_template)
                    bar_code = bar_code + nameImge
                    # Aplicar las imagenes a la hoja
                    if sheet_name in posiciones_imagenes and posiciones_imagenes[sheet_name]:
                        try:
                            sheet_template = wb[sheet_name]
                            aplicar_imagenes_a_hoja(sheet, posiciones_imagenes[sheet_name], sheet_template, start_row)
                        except Exception as e:
                            print(f"Error al procesar imágenes para {sheet_name}: {str(e)}")
                            continue


                    start_row = max_row
        #aplicar formato de las hojas
        sheet_info = info_excel['001'] 
        if 'page_setup' in sheet_info:
            page_setup = sheet_info['page_setup']
            print("page_setup: ", page_setup)
            sheet.page_setup.orientation = page_setup['orientation']
            sheet.page_setup.paperSize = page_setup['paper_size']
            sheet.page_setup.fitToWidth = page_setup['fit_to_width']
            sheet.page_setup.fitToHeight = page_setup['fit_to_height']
            sheet.page_setup.scale = page_setup['scale']
            sheet.page_margins.top = page_setup['margin_top']
            sheet.page_margins.bottom = page_setup['margin_bottom']
            sheet.page_margins.left = page_setup['margin_left']
            sheet.page_margins.right = page_setup['margin_right']
            sheet.print_area = ""
            if sheet.page_setup.fitToWidth is None:
                sheet.page_setup.fitToWidth = 1  # Forzar ajuste a una página de ancho
            if sheet.page_setup.fitToHeight is None:
                sheet.page_setup.fitToHeight = 1  # Forzar ajuste a una página de alto
            if sheet.page_setup.paperSize is None:
                sheet.page_setup.paperSize = sheet.PAPERSIZE_A4 
            
        message = message + "aplicar formato de las hojas exitosamente: "  + "\n"
        if os.path.exists(ruta_report_excel):
            os.remove(ruta_report_excel)

        workbook.save(ruta_report_excel)

        # eliminar images de código de barras
        for nameImge in bar_code:
            os.remove(nameImge)

        message += "Archivo creado exitosamente: " + ruta_report_excel + "\n"
    except Exception as e:
        message += "Error al crear el reporte: " + ruta_template_excel + "\n"
        message += "Error: " + str(e) + "\n"
        log(rout_log, "ser_excel", message)
    return message, ruta_report_excel, principal_sheet.title
