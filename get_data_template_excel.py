from openpyxl import load_workbook

def get_data_template_excel(ruta_excel):
    workbook = load_workbook(ruta_excel)
    info_excel = {}

    for sheet in workbook.worksheets:
        sheet_info = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_info = {
                        'value': cell.value,
                        'font': {
                            'name': cell.font.name,
                            'size': cell.font.size,
                            'bold': cell.font.bold,
                            'italic': cell.font.italic,
                            'underline': cell.font.underline,
                            'color': cell.font.color.rgb if cell.font.color else None
                        },
                        'fill': {
                            'fgColor': cell.fill.fgColor.rgb if cell.fill.fgColor else None,
                            'patternType': cell.fill.patternType
                        },
                        'border': {
                            'left': cell.border.left.style,
                            'right': cell.border.right.style,
                            'top': cell.border.top.style,
                            'bottom': cell.border.bottom.style
                        },
                        'alignment': {
                            'horizontal': cell.alignment.horizontal,
                            'vertical': cell.alignment.vertical,
                            'wrap_text': cell.alignment.wrap_text
                        },
                        'number_format': cell.number_format,
                        'row': cell.row,
                        'column': cell.column,
                        'merge_cells': cell.coordinate in sheet.merged_cells
                    }
                    sheet_info[cell.coordinate] = cell_info

        info_excel[sheet.title] = sheet_info

    return info_excel

